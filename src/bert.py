import numpy as np
import re
from nltk.corpus import stopwords
from nltk import pos_tag
from pathlib import Path
import torch
import json
import os
import openpyxl
from wordfreq import zipf_frequency
from sklearn.metrics.pairwise import cosine_similarity as cosine
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.models import load_model
from transformers import BertTokenizer, BertForMaskedLM

stop_words_ = set(stopwords.words('english'))


def cleaner(word):
    word = re.sub(r'((http|https)\:\/\/)?[a-zA-Z0-9\.\/\?\:@\-_=#]+\.([a-zA-Z]){2,6}([a-zA-Z0-9\.\&\/\?\:@\-_=#])*',
                  '', word, flags=re.MULTILINE)
    word = re.sub('[\W]', ' ', word)
    word = re.sub('[^a-zA-Z]', ' ', word)
    return word.lower().strip()


def process_input(input_text, word2index, sent_max_length):
    input_text = cleaner(input_text)
    clean_text = []
    index_list = []
    input_token = []
    for i, word in enumerate(input_text.split()):
        if word in word2index:
            clean_text.append(word)
            input_token.append(word2index[word])
        else:
            index_list.append(i)
    input_padded = pad_sequences(maxlen=sent_max_length, sequences=[input_token], padding="post", value=0)
    return input_padded, index_list, len(clean_text)


def complete_missing_word(pred_binary, index_list, len_list):
    list_cwi_predictions = list(pred_binary[0][:len_list])
    for i in index_list:
        list_cwi_predictions.insert(i, 0)
    return list_cwi_predictions


def get_bert_candidates(input_text, list_cwi_predictions, tokenizer, model):
    numb_predictions_displayed = 10
    list_candidates_bert = []
    for word, pred in zip(input_text.split(), list_cwi_predictions):
        # if (pred and (pos_tag([word])[0][1] in ['NNS', 'NN', 'VBP', 'RB', 'VBG', 'VBD'])) or (zipf_frequency(word, 'en')) < 3.1:
        if pred or (pos_tag([word])[0][1] in ['NNS', 'NN', 'VBP', 'RB', 'VBG', 'VBD'] and (zipf_frequency(word, 'en')) < 3.1):
            replace_word_mask = input_text.replace(word, '[MASK]')
            text = f'[CLS]{replace_word_mask} [SEP] {input_text} [SEP] '
            tokenized_text = tokenizer.tokenize(text)
            masked_index = [i for i, x in enumerate(tokenized_text) if x == '[MASK]'][0]
            indexed_tokens = tokenizer.convert_tokens_to_ids(tokenized_text)
            segments_ids = [0] * len(tokenized_text)
            tokens_tensor = torch.tensor([indexed_tokens])
            segments_tensors = torch.tensor([segments_ids])
            with torch.no_grad():
                outputs = model(tokens_tensor, token_type_ids=segments_tensors)
                predictions = outputs[0][0][masked_index]
            predicted_ids = torch.argsort(predictions, descending=True)[:numb_predictions_displayed]
            predicted_tokens = tokenizer.convert_ids_to_tokens(list(predicted_ids))
            temp = []
            for sub in predicted_tokens:
                if zipf_frequency(sub, 'en') > 3 and sub != word:
                    temp.append(sub)
            list_candidates_bert.append((word, temp))
    return list_candidates_bert


def getWordmap(wordVecPath):
    words = []
    We = []
    f = open(wordVecPath, 'r')
    lines = f.readlines()

    for (n, line) in enumerate(lines):
        word, vect = line.rstrip().split(' ', 1)
        vect = np.fromstring(vect, sep=' ')
        We.append(vect)
        words.append(word)
    f.close()
    return words, We


def getWordCount(word_count_path):
    word2count = {}
    xlsx_file = Path('', word_count_path)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    last_column = sheet.max_column - 1
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        word2count[row[0]] = round(float(row[last_column]), 3)

    return word2count


def preprocess_SR(source_word, candidate_words, fasttext_dico, fasttext_emb, word_count):
    ss = []
    sis_scores = []
    count_scores = []

    isFast = True

    if source_word not in fasttext_dico:
        isFast = False
    else:
        source_emb = fasttext_emb[fasttext_dico.index(source_word)].reshape(1, -1)

    for sub in candidate_words:

        if sub not in word_count:
            continue
        else:
            sub_count = word_count[sub]

        if isFast:
            if sub not in fasttext_dico:
                continue

            token_index_fast = fasttext_dico.index(sub)
            sis = cosine(source_emb, fasttext_emb[token_index_fast].reshape(1, -1))

            sis_scores.append(sis)

        ss.append(sub)
        count_scores.append(sub_count)
    return ss, sis_scores, count_scores


def compute_context_sis_score(source_word, sis_context, substitution_selection, fasttext_dico, fasttext_emb):
    context_sis = []

    word_context = []

    for con in sis_context:
        if con == source_word or (con not in fasttext_dico):
            continue

        word_context.append(con)

    if len(word_context) != 0:
        for sub in substitution_selection:
            sub_emb = fasttext_emb[fasttext_dico.index(sub)].reshape(1, -1)
            all_sis = 0
            for con in word_context:
                token_index_fast = fasttext_dico.index(con)
                all_sis += cosine(sub_emb, fasttext_emb[token_index_fast].reshape(1, -1))

            context_sis.append(all_sis / len(word_context))
    else:
        for i in range(len(substitution_selection)):
            context_sis.append(len(substitution_selection) - i)

    return context_sis


def substitution_ranking(source_word, candidate_words, fasttext_dico, fasttext_emb, word_count):
    ss, sis_scores, count_scores = preprocess_SR(source_word, candidate_words, fasttext_dico, fasttext_emb, word_count)

    if len(ss) == 0:
        return source_word

    if len(sis_scores) > 0:
        seq = sorted(sis_scores, reverse=True)
        sis_rank = [seq.index(v) + 1 for v in sis_scores]

    rank_count = sorted(count_scores, reverse=True)

    count_rank = [rank_count.index(v) + 1 for v in count_scores]

    bert_rank = []
    for i in range(len(ss)):
        bert_rank.append(i + 1)

    if len(sis_scores) > 0:
        all_ranks = [bert + sis + count for bert, sis, count in zip(bert_rank, sis_rank, count_rank)]
    else:
        all_ranks = [bert + count for bert, count in zip(bert_rank, count_rank)]

    pre_index = all_ranks.index(min(all_ranks))
    pre_word = ss[pre_index]

    return pre_word


def run(list_texts: list):
    model_name = 'model_CWI.h5'
    path_dir = os.getcwd() + f"/model/{model_name}"
    model_CWI = load_model(path_dir)
    with open(os.getcwd() + '/dataset/word2index.json', 'r') as instream:
        word2index = json.load(instream)
    with open(os.getcwd() + '/dataset/sent_max_length.txt', 'r') as instream:
        sent_max_length = int(instream.readline().strip())

    bert_model = 'bert-large-uncased'
    tokenizer = BertTokenizer.from_pretrained(bert_model)
    model = BertForMaskedLM.from_pretrained(bert_model)

    fasttext_dico, fasttext_emb = getWordmap(os.getcwd() + '/model/embeddings/glove.6B.300d.txt')
    word_count = getWordCount(os.getcwd() + '/dataset/SUBTLEX_frequency.xlsx')
    newlist = []
    for input_text in list_texts:
        new_text = input_text
        input_padded, index_list, len_list = process_input(input_text, word2index, sent_max_length)
        pred_cwi = model_CWI.predict(input_padded)
        pred_cwi_binary = np.argmax(pred_cwi, axis=2)
        complete_cwi_predictions = complete_missing_word(pred_cwi_binary, index_list, len_list)
        bert_candidates = get_bert_candidates(cleaner(input_text), complete_cwi_predictions, tokenizer, model)
        for word_to_replace, l_candidates in bert_candidates:
            pre_word = substitution_ranking(word_to_replace, l_candidates, fasttext_dico, fasttext_emb, word_count)
        #     tuples_word_zipf = []
        #     for w in l_candidates:
        #         if w.isalpha():
        #             tuples_word_zipf.append((w, zipf_frequency(w, 'en')))
        #     tuples_word_zipf = sorted(tuples_word_zipf, key=lambda x: x[1], reverse=True)
        #     if len(tuples_word_zipf) != 0:
            new_text = re.sub(word_to_replace, pre_word, new_text)
        new_text = re.sub('-LRB-', '(', new_text)
        new_text = re.sub('-RRB-', ')', new_text)
        newlist.append(new_text)
    return newlist


if __name__ == '__main__':
    run(['Hurwicz, who had graduated from Warsaw University in 1938, at the time of Nazi invasion on Poland was in London, moved to Switzerland then to Portugal and finally in 1940 he emigrated to the United States.'])